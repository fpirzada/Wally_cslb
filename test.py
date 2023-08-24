# # #     MainContent_BusInfo = soup.find(id='MainContent_BusInfo')
# #     #     if MainContent_BusInfo:
# #     #         print("-------- Business Information --------")        
# #     #         MainContent_BusInfo = MainContent_BusInfo.get_text()
# #     #         print(f"MainContent_BusInfo : {MainContent_BusInfo}")
# #     #     else:
# #     #         print("************** NO MainContent_BusInfo")

# #     #     MainContent_Entity = soup.find(id='MainContent_Entity')
# #     #     if MainContent_Entity:
# #     #         MainContent_Entity = MainContent_Entity.get_text()
# #     #         print(f"MainContent_Entity : {MainContent_Entity}")
# #     #     else:
# #     #         print("************** NO MainContent_Entity")

# #     #     MainContent_IssDt = soup.find(id='MainContent_IssDt')
# #     #     if MainContent_IssDt:
# #     #         MainContent_IssDt = MainContent_IssDt.get_text()
# #     #         print(f"MainContent_IssDt : {MainContent_IssDt}")
# #     #     else:
# #     #         print("************** NO MainContent_IssDt")

# #     #     MainContent_ExpDt = soup.find(id='MainContent_ExpDt')
# #     #     if MainContent_ExpDt:
# #     #         MainContent_ExpDt = MainContent_ExpDt.get_text()
# #     #         print(f"MainContent_ExpDt : {MainContent_ExpDt}")
# #     #     else:
# #     #         print("************** NO MainContent_ExpDt")

# #     # except Exception as e:
# #     #     print("-------- NO Business Information --------")
# #     #     print(e)
    
# #     # wait = WebDriverWait(driver,2)

# #     # try:
# #     #     MainContent_Status = soup.find(id='MainContent_Status')
# #     #     if MainContent_Status:
# #     #         print("-------- License Status --------")
# #     #         MainContent_Status = MainContent_Status.get_text()
# #     #         print(f"MainContent_Status : {MainContent_Status}")
# #     #     else:
# #     #         print("************** NO MainContent_Status")

# #     #     try:
# #     #         MainContent_AddLicStatus = soup.find(id='MainContent_AddLicStatus')
# #     #         if MainContent_AddLicStatus:
# #     #             print("-------- Additional Status --------")
# #     #             MainContent_AddLicStatus = MainContent_AddLicStatus.get_text()
# #     #             print(f"MainContent_AddLicStatus : {MainContent_AddLicStatus}")
# #     #         else:
# #     #             print("~~ in else Additional Status ~~")
# #     #             print("************** NO MainContent_AddLicStatus")

# #     #     except Exception as e:
# #     #         print("-------- NO Additional Status --------")

# #     # except Exception as e:
# #     #     print("-------- NO License Status --------")
# #     #     print(e)
    
# #     # try:
# #     #     try:
# #     #         MainContent_ClassificationRow2 = soup.find(id='MainContent_ClassificationRow2')
# #     #         if MainContent_ClassificationRow2:
# #     #             print("-------- Classifications --------")
# #     #             elements_by_css = MainContent_ClassificationRow2.select('ul.list-understated li')
# #     #             if elements_by_css:
# #     #                 for element in elements_by_css:
# #     #                     try:
# #     #                         print(f"MainContent_ClassificationRow2_TEXT : {element.get_text()}")
# #     #                         element_a = element.find('a')
# #     #                         print(f"MainContent_ClassificationRow2_HREF : 'https://www.cslb.ca.gov/'+{element_a['href']}")
# #     #                     except Exception as e:
# #     #                         print("~~ in else Classifications ~~")
# #     #                         print(e)
# #     #             else:
# #     #                 print("~~ in else Classifications ~~")
# #     #                 print(f"MainContent_ClassificationRow2_TEXT : {MainContent_ClassificationRow2.get_text()}")
# #     #                 print(f"MainContent_ClassificationRow2_HREF : {MainContent_ClassificationRow2.find('a')['href']}")
# #     #         else:
# #     #             print("************** NO Classifications")
# #     #     except Exception as e:
# #     #         print(e)

        
# #     #     try:
# #     #         MainContent_CertificationRow2 = soup.find(id='MainContent_CertificationRow2')
# #     #         if MainContent_CertificationRow2:
# #     #             print("-------- Certification --------")
# #     #             elements_by_css = MainContent_CertificationRow2.select('ul.list-understated li')
# #     #             if elements_by_css:
# #     #                 for element in elements_by_css:
# #     #                     print(f"MainContent_CertificationRow2_TEXT : {element.get_text()}")
# #     #                     print(f"MainContent_CertificationRow2_HREF : 'https://www.cslb.ca.gov/'+{element_a.find('a')['href']}")
# #     #             else:
# #     #                 print("~~ in else Certification ~~")
# #     #                 print(f"MainContent_CertificationRow2_TEXT : {MainContent_CertificationRow2.get_text()}")
# #     #                 print(f"MainContent_CertificationRow2_HREF : {MainContent_CertificationRow2.find('a')['href']}")
# #     #         else:
# #     #             print("~~ in else Certification ~~")
# #     #             print("************** NO Certification")

# #     #     except Exception as e:
# #     #         print("-------- NO Certification --------")
# #     #         print(e)

# #     # except Exception as e:
# #     #     print("-------- NO Classifications --------")
# #     #     print(e)

# #     # try:
# #     #     MainContent_BondingCellTable = soup.find(id='MainContent_BondingCellTable')
# #     #     print(MainContent_BondingCellTable)
# #     #     MainContent_BondingCellTable_all = MainContent_BondingCellTable.find_all(['p', 'a', 'table'])
# #     #     print("-------- Bonding Information --------")

# #     #     for element in MainContent_BondingCellTable_all:
# #     #         if element.name == 'table':
# #     #             print("-------- "+element.text+" --------")
# #     #         elif element.name == 'p':
# #     #             print(element.text)
# #     #         elif element.name == 'a':
# #     #             print('Link:', element['href'])


# #         # print(soup.find(id='MainContent_LicTable'))


# #         # MainContent_BondingCellTable_header = MainContent_BondingCellTable.select('.CenterAlign')
# #         # MainContent_BondingCellTable_p = MainContent_BondingCellTable.find_all('p')
# #         # if MainContent_BondingCellTable:
# #         #     print("-------- Contractor's Bond --------")
# #         #     print(MainContent_BondingCellTable_header[0].get_text())
# #         #     print(MainContent_BondingCellTable_p[0].get_text())
# #         #     print(MainContent_BondingCellTable_p[1].get_text())
# #         #     print(MainContent_BondingCellTable_p[2].get_text())
# #         #     print(MainContent_BondingCellTable_p[3].get_text())
            
# #         #     try:
# #         #         MainContent_BondingCellTable_a = MainContent_BondingCellTable.find_all('a')
# #         #         for a in MainContent_BondingCellTable_a:
# #         #             if "Contractor's" in a.get_text():
# #         #                 print("https://www.cslb.ca.gov/"+a['href'])
# #         #     except Exception as e:
# #         #         print("************** NO Contractor's Bond History")

# #         # try:
# #         #     print("-------- Bond of Qualifying Individual --------")
# #         #     print(MainContent_BondingCellTable_header[1].get_text())
# #         #     print(MainContent_BondingCellTable_p[4].get_text())
# #         #     print(MainContent_BondingCellTable_p[5].get_text())
# #         #     try:
# #         #         for a in MainContent_BondingCellTable_a:
# #         #             if "BQI's" in a.get_text():
# #         #                 print(a['href'])
# #         #     except Exception as e:
# #         #         print("************** NO BQI's Bond History")
# #         # except Exception as e:
# #         #     print("-------- NO Bond of Qualifying Individual --------")


# # from bs4 import BeautifulSoup

# # from selenium.webdriver.common.by import By
# # from selenium.webdriver.support import expected_conditions as EC
# # from selenium.webdriver.support.wait import WebDriverWait
# # from selenium.webdriver.support.ui import Select
# # from selenium.webdriver.common.keys import Keys
# # from selenium.webdriver.common.action_chains import ActionChains


# # data  = """<table id="MainContent_dlAssociated" cellspacing="0" style="border-collapse:collapse;">
# # 		<tbody><tr>
# # 			<td>
# #                 <table>
# #                     <tbody><tr>
# #                         <td class="tblHeader">
# #                             <span id="MainContent_dlAssociated_lblName_0">Name</span>
# #                         </td>
# #                         <td>
# #                             <a id="MainContent_dlAssociated_hlName_0" href="/OnlineServices/CheckLicenseII/PersonnelDetail.aspx?LicNum=855764&amp;SeqNumber=892728&amp;LicName=U1BFRURXQVkrQ09OU1RSVUNUT1JTK0lOQw==">WILLIAM JAVIER FORERO </a>
# #                         </td>
# #                     </tr>
# #                     <tr>
# #                         <td class="tblHeader">
# #                             <span id="MainContent_dlAssociated_lblTitletxt_0">Title</span>
# #                         </td>
# #                         <td>
# #                             <span id="MainContent_dlAssociated_lblTitle_0">RMO / CEO / PRES</span>

# #                         </td>
# #                     </tr>
# #                     <tr>
# #                         <td class="tblHeader">
# #                             <span id="MainContent_dlAssociated_lblDate_0">Association Date</span>
# #                         </td>
# #                         <td>
# #                             <span id="MainContent_dlAssociated_lblAssociationDate_0">05/10/2010</span>
# #                         </td>
# #                     </tr>
# #                     <tr>
# #                         <td class="tblHeader">
# #                             <span id="MainContent_dlAssociated_lbltxtClassification_0">Classification</span>
# #                         </td>
# #                         <td>
# #                             <span id="MainContent_dlAssociated_lblClassification_0">B</span>
# #                         </td>
# #                     </tr>
# #                     <tr>
# #                         <td class="tblHeader">
# #                             <span id="MainContent_dlAssociated_lblAClassification_0">Additional Classification</span>
# #                         </td>
# #                         <td>
# #                             <a id="MainContent_dlAssociated_hlAdditionalClassification_0" href="/OnlineServices/CheckLicenseII/PersonnelDetail.aspx?LicNum=855764&amp;SeqNumber=892728&amp;LicName=U1BFRURXQVkrQ09OU1RSVUNUT1JTK0lOQw==">There are additional classifications that can be viewed by selecting this link.</a>
# #                         </td>
# #                     </tr>
# #                 </tbody></table>
# #             </td>
# # 		</tr>
# # 	</tbody></table>
# # """


# # # Parse the HTML using Beautiful Soup
# # soup = BeautifulSoup(data, 'html.parser')

# # soup_ = soup.find(id="MainContent_dlAssociated").find_all("table")
# # for _ in soup_:
# #     tr = _.find_all("tr")
# #     for tr_ in tr:
# #         print("--")
# #         print(tr_.text)


# # # # print(soup[0])
# # # soup = soup.find_all('tr')

# # # for tr in soup:
# # #     print("--")
# # #     if tr.select('h2.subheading'):
# # #         print("-------- "+tr.text+" --------")
# # #     if tr.select('td#MainContent_BusInfo'):
# # #         print(list(tr.stripped_strings))
# # #     if tr.select("td#MainContent_Entity"):
# # #         print("Entity : "+tr.text)
# # #     if tr.select("td#MainContent_IssDt"):
# # #         print("Issue Date : "+tr.text)
# # #     if tr.select("td#MainContent_ReIssueDt"):
# # #         print("Reissue Date : "+tr.text)
# # #     if tr.select("td#MainContent_ExpDt"):
# # #         print("Expire Date : "+tr.text)

    
# # #     if tr.select("td#MainContent_Status"):
# # #         print(tr.text)
# # #     if tr.select("td#MainContent_ClassCellTable"):
# # #         a_ = tr.find_all('a')
# # #         for a in a_:
# # #             print(a.text)
# # #             print(a['href'])

# # #     if tr.select("td#MainContent_BondingCellTable"):
# # #         MainContent_BondingCellTable_all = tr.find_all(['p', 'a', 'table'])
# # #         for element in MainContent_BondingCellTable_all:
# # #             if element.name == 'table':
# # #                 print("-------- "+element.text+" --------")
# # #             elif element.name == 'p':
# # #                 print(element.text)
# # #             elif element.name == 'a':
# # #                 print('Link:', element['href'])

# # #     if tr.select("td#MainContent_WCStatus"):
# # #         print(list(tr.stripped_strings))
# # #         if tr.find("a"):
# # #             print(tr.find("a")['href'])

# # #     if tr.select("td#MainContent_LLIStatus"):
# # #         MainContent_LLIStatus_all = tr.find_all(['p'])
# # #         for element in MainContent_LLIStatus_all:
# # #             if element.name == 'p':
# # #                 print(element.text)

# # #     if tr.select("td#MainContent_ActionCodesCellTable"):
# # #         MainContent_ActionCodesCellTable_all = tr.find_all(['li'])
# # #         for element in MainContent_ActionCodesCellTable_all:
# # #             if element.name == 'li':
# # #                 print(element.text)

# # #     if tr.select("td#MainContent_MultiLicDisplay"):
# # #         print(tr.text)
    
# #     #  print(tr)
# # # xx = soup.find_all(['p', 'a', 'table'])
# # # for element in xx:
# # #         if element.name == 'table':
# # #             if 'class="subheading"' in element:
# # #                 print("-------- "+element.text+" --------")
# # #             else:
# # #                 print("-------- "+element.text+" --------")

# # #         elif element.name == 'p':
# # #             print(element.text)
# # #         elif element.name == 'a':
# # #             print('Link:', element['href'])
            

# # # # Use CSS selector to find elements with a specific class
# # # elements_with_class = soup.find_all('a')
# # # print(len(elements_with_class))
# # # print(elements_with_class[0].get_text())
# # # print(elements_with_class[0]['href'])




# # import openpyxl
# # from openpyxl.styles import Alignment

# # # Create a new Excel workbook and select the active worksheet
# # workbook = openpyxl.Workbook()
# # worksheet = workbook.active

# # # Define the top headers and their corresponding subheaders
# # header_subheader_mapping = {
# #     "Top Header 1": ["Sub Header 1.1", "Sub Header 1.2", "Sub Header 1.3"],
# #     "Top Header 2": ["Sub Header 2.1", "Sub Header 2.2"],
# #     "Top Header 3": ["Sub Header 3.1"]
# # }

# # # Merge cells for the headers and format them
# # col_offset = 1
# # for top_header, subheaders in header_subheader_mapping.items():
# #     col_num = col_offset
# #     worksheet.cell(row=1, column=col_num, value=top_header)
# #     worksheet.merge_cells(start_row=1, start_column=col_num, end_row=1, end_column=col_num + len(subheaders) - 1)
# #     worksheet.cell(row=1, column=col_num).alignment = Alignment(horizontal='center', vertical='center')

# #     for subheader in subheaders:
# #         worksheet.cell(row=2, column=col_num, value=subheader)
# #         col_num += 1

# #     col_offset += len(subheaders)

# # # Insert an empty row as the 3rd row
# # worksheet.insert_rows(3)

# # # Sample data
# # data = [
# #     [],  # Empty row
# #     ["Data 1.1", "Data 1.2", "Data 1.3", "Data 2.1", "Data 2.2", "Data 3.1"],
# #     ["Data 4.1", "Data 4.2", "Data 4.3", "Data 5.1", "Data 5.2", "Data 6.1"],
# #     ["Data 7.1", "Data 7.2", "Data 7.3", "Data 8.1", "Data 8.2", "Data 9.1"]
# # ]

# # # Place data under each subheader
# # for row_data in data:
# #     worksheet.append(row_data)

# # # Save the Excel file
# # workbook.save("output.xlsx")




# # import openpyxl
# # from openpyxl.styles import Alignment
# # import os
# # import datetime

# # def create_or_load_workbook(filename):
# #     if os.path.exists(filename):
# #         return openpyxl.load_workbook(filename)
# #     else:
# #         workbook = openpyxl.Workbook()
# #         workbook.save(filename)
# #         return openpyxl.load_workbook(filename)

# # def format_header_and_subheader(worksheet, header_subheader_mapping):
# #     col_offset = 1
# #     for top_header, subheaders in header_subheader_mapping.items():
# #         col_num = col_offset
# #         worksheet.cell(row=1, column=col_num, value=top_header)
# #         worksheet.merge_cells(start_row=1, start_column=col_num, end_row=1, end_column=col_num + len(subheaders) - 1)
# #         worksheet.cell(row=1, column=col_num).alignment = Alignment(horizontal='center', vertical='center')

# #         for subheader in subheaders:
# #             worksheet.cell(row=2, column=col_num, value=subheader)
# #             col_num += 1

# #         col_offset += len(subheaders)

# #     # Insert an empty row as the 3rd row
# #     worksheet.insert_rows(3)

# # def append_live_data(worksheet, live_data):
# #     for row_data in live_data:
# #         worksheet.append(row_data)

# # def auto_adjust_column_widths(worksheet):
# #     for col_idx, column_cells in enumerate(worksheet.columns, start=1):
# #         max_length = max(len(str(cell.value)) for cell in column_cells)
# #         col_letter = openpyxl.utils.get_column_letter(col_idx)
# #         adjusted_width = max(max_length, len(col_letter)) + 2
# #         worksheet.column_dimensions[col_letter].width = adjusted_width

# # def save_workbook(workbook, filename):
# #     workbook.save(filename)

# # def main():
# #     current_date = datetime.datetime.now().strftime("%Y-%m-%d")
# #     filename = f"output_{current_date}.xlsx"
    
# #     workbook = create_or_load_workbook(filename)
# #     worksheet = workbook.active

# #     header_subheader_mapping = {
# #         "Top Header 1": ["Sub Header 1.1", "Sub Header 1.2", "Sub Header 1.3"],
# #         "Top Header 2": ["Sub Header 2.1", "Sub Header 2.2"],
# #         "Top Header 3": ["Sub Header 3.1"]
# #     }

# #     format_header_and_subheader(worksheet, header_subheader_mapping)

# #     live_data = [
# #         ["Live Data 1.1", "Live Data 1.2", "Live Data 1.3", "Live Data 2.1", "Live Data 2.2", "Live Data 3.1"],
# #         ["Live Data 4.1", "Live Data 4.2", "Live Data 4.3", "Live Data 5.1", "Live Data 5.2", "Live Data 6.1"]
# #     ]

# #     append_live_data(worksheet, live_data)

# #     auto_adjust_column_widths(worksheet)

# #     save_workbook(workbook, filename)

# # if __name__ == "__main__":
# #     main()





# import openpyxl
# from openpyxl.styles import Alignment
# import os
# import datetime

# def create_or_load_workbook(filename):
#     if os.path.exists(filename):
#         return openpyxl.load_workbook(filename)
#     else:
#         workbook = openpyxl.Workbook()
#         workbook.save(filename)
#         return openpyxl.load_workbook(filename)

# def format_header_and_subheader(worksheet, header_subheader_mapping):
#     col_offset = 1
#     for top_header, subheaders in header_subheader_mapping.items():
#         col_num = col_offset
#         worksheet.cell(row=1, column=col_num, value=top_header)
#         worksheet.merge_cells(start_row=1, start_column=col_num, end_row=1, end_column=col_num + len(subheaders) - 1)
#         worksheet.cell(row=1, column=col_num).alignment = Alignment(horizontal='center', vertical='center')

#         for subheader in subheaders:
#             worksheet.cell(row=2, column=col_num, value=subheader)
#             col_num += 1

#         col_offset += len(subheaders)

#     # Insert an empty row as the 3rd row
#     worksheet.insert_rows(3)

# def append_live_data(worksheet, live_data):
#     for row_data in live_data:
#         worksheet.append(row_data)

# def auto_adjust_column_widths(worksheet):
#     for col_idx, column_cells in enumerate(worksheet.columns, start=1):
#         max_length = max(len(str(cell.value)) for cell in column_cells)
#         col_letter = openpyxl.utils.get_column_letter(col_idx)
#         adjusted_width = max(max_length, len(col_letter)) + 2
#         worksheet.column_dimensions[col_letter].width = adjusted_width

# def save_workbook(workbook, filename):
#     workbook.save(filename)

# def main(header_subheader_mapping,live_data):
#     current_date = datetime.datetime.now().strftime("%Y-%m-%d")
#     filename = f"output_{current_date}.xlsx"
    
#     workbook = create_or_load_workbook(filename)
#     worksheet = workbook.active

    

#     format_header_and_subheader(worksheet, header_subheader_mapping)

    

#     append_live_data(worksheet, live_data)

#     auto_adjust_column_widths(worksheet)

#     save_workbook(workbook, filename)

# if __name__ == "__main__":
#     header_subheader_mapping = {
#         "Business Information": ["License #", "Name", "Address","City","Zip","Phone #","Entity","Issue Date","Reissue Date","Expire Date"],
#         "License Status": ["Status"],
#         "Classifications": ["Classifications Name","Certifications Name"],
#         "Bonding Information": ["Contractor's Bond Title","Contractor's Bond Number ","Contractor's Bond Amount","Contractor's Bond Effective Date","Contractor's Bond  History Link","LLC EMPLOYEE/WORKER BOND Title","LLC EMPLOYEE/WORKER BOND Number","LLC EMPLOYEE/WORKER BOND Amount","LLC EMPLOYEE/WORKER BOND Effective Date","LLC EMPLOYEE/WORKER BOND History Link","Bond of Qualifying Individual","Bond of Qualifying Individual Title","Bond of Qualifying Individual Effective Date","Bond of Qualifying Individual History Link"],
#         "Workers' Compensation": ["Title","Policy Number","Effective Date","Expire Date","Workers' Compensation History Link","Insurance Company Code",],
#         "Liability Insurance Information": ["Title","Policy Number","Amount","Effective Date","Expiration Date","Liability Insurance History","Insurance Company"],
#         "Miscellaneous Information": ["Title"],
#         "Other": ["Title"],
#         "Personnel": ["Name","Title","Association Date",""],

#     }
#     live_data = [
#         ["Live Data 1.1", "Live Data 1.2", "Live Data 1.3", "Live Data 2.1", "Live Data 2.2", "Live Data 3.1"],
#         ["Live Data 4.1", "Live Data 4.2", "Live Data 4.3", "Live Data 5.1", "Live Data 5.2", "Live Data 6.1"]
#     ]
#     main(header_subheader_mapping,live_data)



# """"""


# import pandas as pd

# # Paths
# template_file_path = "temp/temp.xlsx"
# output_file_path = "output.xlsx"  # Adjust the output file name as needed

# # Data to append
# new_data = [
#     ["Business A", "License Class A", "Class A Name", "example.com", "email@example.com", "CSLB Link A",
#      "123 Main St", "City A", "State A", "12345", "License A", "123-456-7890", "PEO Code A", "2023-08-24",
#      "PEO Name A", "Category A", "2023-08-24", "Category old A", "Policy Number A", "2023-12-31",
#      "Campaign A", 10, "John", "Doe", "Manager", "john.doe@example.com", "linkedin.com/johndoe", "LinkedIn Company A"]
#     # Add more rows as needed
# ]

# # Read the template file
# template_df = pd.read_excel(template_file_path)

# # Create a DataFrame from the new data
# new_data_df = pd.DataFrame(new_data, columns=template_df.columns)

# # Concatenate the DataFrames
# updated_df = pd.concat([template_df, new_data_df], ignore_index=True)

# # Save the updated DataFrame to the output file
# updated_df.to_excel(output_file_path, index=False)

# print(f"Data appended and saved to '{output_file_path}'")



# from openpyxl import load_workbook
# import pandas as pd

# # Paths
# template_file_path = "temp/temp.xlsx"
# output_file_path = "output.xlsx"  # Adjust the output file name as needed

# # Data to append
# new_data = [
#     ["Business A", "License Class A", "Class A Name", "example.com", "email@example.com", "CSLB Link A",
#      "123 Main St", "City A", "State A", "12345", "License A", "123-456-7890", "PEO Code A", "2023-08-24",
#      "PEO Name A", "Category A", "2023-08-24", "Category old A", "Policy Number A", "2023-12-31",
#      "Campaign A", 10, "John", "Doe", "Manager", "john.doe@example.com", "linkedin.com/johndoe", "LinkedIn Company A"]
#     # Add more rows as needed
# ]

# # Load the template Excel file using openpyxl
# template_wb = load_workbook(template_file_path)
# template_sheet = template_wb.active

# # Insert rows for new data at the second row
# for data_row in new_data:
#     template_sheet.insert_rows(2)  # Insert a row above the current second row
#     for col, value in enumerate(data_row, start=1):
#         template_sheet.cell(row=2, column=col, value=value)

# # Save the modified workbook to the output file
# template_wb.save(output_file_path)

# print(f"Data appended and saved to '{output_file_path}'")





# import xlwings as xw

# # Paths
# template_file_path = "temp/temp.xlsx"
# output_file_path = "output.xlsx"  # Adjust the output file name as needed

# # Data to append
# new_data = [
#     ["Business A", "License Class A", "Class A Name", "example.com", "email@example.com", "CSLB Link A",
#      "123 Main St", "City A", "State A", "12345", "License A", "123-456-7890", "PEO Code A", "2023-08-24",
#      "PEO Name A", "Category A", "2023-08-24", "Category old A", "Policy Number A", "2023-12-31",
#      "Campaign A", 10, "John", "Doe", "Manager", "john.doe@example.com", "linkedin.com/johndoe", "LinkedIn Company A"]
#     # Add more rows as needed
# ]

# # Load the Excel template using xlwings
# template_wb = xw.Book(template_file_path)
# template_sheet = template_wb.sheets[0]

# # Define the column indices where you want to insert data
# column_indices = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26]

# # Insert rows for new data at the second row
# template_sheet.api.Rows("2:2").insert()

# # Write new data to specific columns in the second row
# for idx, value in zip(column_indices, new_data[0]):
#     template_sheet.cells(2, idx + 1).value = value

# # Save the modified workbook to the output file
# template_wb.save(output_file_path)
# template_wb.close()

# print(f"Data appended and saved to '{output_file_path}'")


# from openpyxl import load_workbook

# # Paths
# template_file_path = "output.xlsx"  # Replace with the actual output file path

# # Load the Excel file using openpyxl
# wb = load_workbook(template_file_path)
# sheet = wb.active

# # Find and remove empty rows
# rows_to_delete = []
# for row in range(sheet.max_row, 1, -1):  # Iterate in reverse order
#     empty_row = all(cell.value is None for cell in sheet[row])
#     if empty_row:
#         rows_to_delete.append(row)

# for row in rows_to_delete:
#     sheet.delete_rows(row)

# # Save the modified workbook
# wb.save(template_file_path)
# wb.close()

# print(f"Empty rows removed and file saved: '{template_file_path}'")






# from openpyxl import load_workbook

# # Path to the Excel file
# output_file_path = "output.xlsx"  # Adjust the output file name as needed

# # Load the existing data from the output file if it exists
# try:
#     existing_wb = load_workbook(output_file_path)
#     existing_sheet = existing_wb.active
# except FileNotFoundError:
#     existing_wb = None

# if existing_wb:
#     empty_row = None
#     for row_idx, row in enumerate(existing_sheet.iter_rows(values_only=True), start=1):
#         if all(cell is None for cell in row):
#             empty_row = row_idx
#             break
    
#     if empty_row is None:
#         empty_row = existing_sheet.max_row + 1
# else:
#     empty_row = 1  # Start from the first row if the file doesn't exist

# print(f"The first empty row is: {empty_row}")



# from openpyxl import load_workbook
# from openpyxl.utils.dataframe import dataframe_to_rows
# from openpyxl.styles import PatternFill
# from datetime import datetime

# # Paths
# template_file_path = "temp/temp.xlsx"
# output_file_path = "output.xlsx"  # Adjust the output file name as needed

# # Data to append
# new_data = [
#     ["Business A", "License Class A", "Class A Name", "example.com", "email@example.com", "CSLB Link A",
#      "123 Main St", "City A", "State A", "12345", "License A", "123-456-7890", "PEO Code A", "2023-08-24",
#      "PEO Name A", "Category A", "2023-08-24", "Category old A", "Policy Number A", "2023-12-31",
#      "Campaign A", 10, "John", "Doe", "Manager", "john.doe@example.com", "linkedin.com/johndoe", "LinkedIn Company A"]
#     # Add more rows as needed
# ]

# # Load the template Excel file using openpyxl
# template_wb = load_workbook(template_file_path)
# template_sheet = template_wb.active

# # Insert rows for new data at the second row
# # empty_row = template_sheet.max_row + 1

# template_sheet.cell(row=empty_row, column=2, value="dddd")


# # Write new data to specific columns in the second row
# # for col_idx, value in enumerate(new_data[0], start=1):
#     # template_sheet.cell(row=2, column=col_idx, value=value)

# # Adjust formatting if needed (example: fill with yellow)
# # yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
# # for row in template_sheet.iter_rows(min_row=2, max_row=2, min_col=1, max_col=template_sheet.max_column):
# #     for cell in row:
# #         cell.fill = yellow_fill

# # Save the modified workbook to the output file
# template_wb.save(output_file_path)

# print(f"Data appended and saved to '{output_file_path}'")




################## working


# from openpyxl import load_workbook

# try:
#     # Paths
#     output_file_path = "output.xlsx"  # Adjust the output file name as needed
#     existing_wb = load_workbook(output_file_path)
#     existing_sheet = existing_wb.active
# except:
#     template_file_path = "temp/temp.xlsx"
#     # Load the template Excel file using openpyxl
#     existing_wb = load_workbook(template_file_path)
#     existing_sheet = existing_wb.active

# # Data for the new record
# new_record_data = [
#     "Business B", "License Class B", "Class B Name",  # Add more data fields as needed
# ]

# if existing_wb:
#     empty_row = None
#     for row_idx, row in enumerate(existing_sheet.iter_rows(values_only=True), start=1):
#         if all(cell is None for cell in row):
#             empty_row = row_idx
#             break
    
#     if empty_row is None:
#         empty_row = existing_sheet.max_row + 1
# else:
#     empty_row = 1  # Start from the first row if the file doesn't exist

# # Add the new record data to the empty row
# for col_idx, value in enumerate(new_record_data, start=1):
#     existing_sheet.cell(row=empty_row, column=col_idx, value=value)

# # Save the modified workbook to the output file
# if existing_wb:
#     existing_wb.save(output_file_path)
# else:
#     existing_wb = openpyxl.Workbook()
#     existing_wb.active = existing_wb.active  # Ensure there's at least one sheet
#     existing_wb.save(output_file_path)

# print(f"New record appended to row {empty_row} in '{output_file_path}'")




from openpyxl import load_workbook

def load_existing_workbook(file_path):
    try:
        existing_wb = load_workbook(file_path)
        existing_sheet = existing_wb.active
    except FileNotFoundError:
        existing_wb = None
        existing_sheet = None
    return existing_wb, existing_sheet

def find_empty_row(sheet):
    empty_row = None
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if all(cell is None for cell in row):
            empty_row = row_idx
            break

    if empty_row is None:
        empty_row = sheet.max_row + 1
    return empty_row

def append_new_record(sheet, empty_row, new_record_data,col_idx_value):
    # for col_idx, value in enumerate(new_record_data, start=1):
    for col_idx,value in col_idx_value:
        sheet.cell(row=empty_row, column=col_idx, value=value)




def run(col_idx_value):
    # Paths
    output_file_path = "output.xlsx"  # Adjust the output file name as needed
    template_file_path = "temp/temp.xlsx"

    # Load or create the workbook
    existing_wb, existing_sheet = load_existing_workbook(output_file_path)
    if existing_wb is None:
        existing_wb = load_workbook(template_file_path)
        existing_sheet = existing_wb.active

    # Data for the new record
    new_record_data = [
        "Business B", "License Class B", "Class B Name",  # Add more data fields as needed
    ]

    # Find an empty row to append the new record
    empty_row = find_empty_row(existing_sheet)

    # Append the new record data to the empty row
    append_new_record(existing_sheet, empty_row, new_record_data,col_idx_value)

    # Save the modified workbook to the output file
    existing_wb.save(output_file_path)

    print(f"New record appended to row {empty_row} in '{output_file_path}'")

if __name__ == "__main__":
    col_idx_value = [(1, 'MARCAURELLE ENGINEERING & CONSTRUCTION'), (6, 'https://www.cslb.ca.gov/OnlineServices/CheckLicenseII/LicenseDetail.aspx?LicNum=804875'), (7, '1551 SHELTER ISLAND DR'), (8, 'SAN DIEGO'), (10, '92106'), (11, '804875'), (12, '(619) 805-6218'), (13, None), (16, 'EXEMPT'), (23, 'WILLIAM'), (24, 'GABRIELSON'), (25, 'RMO / CEO / PRES'), (2, 'B'), (14, '03/03/2022'), (20, 'None')]
    run(col_idx_value)





    ############
    ############



from openpyxl import load_workbook

def load_existing_workbook(file_path):
    try:
        existing_wb = load_workbook(file_path)
        existing_sheet = existing_wb.active
    except FileNotFoundError:
        existing_wb = None
        existing_sheet = None
    return existing_wb, existing_sheet

def find_empty_row(sheet):
    empty_row = None
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if all(cell is None for cell in row):
            empty_row = row_idx
            break

    if empty_row is None:
        empty_row = sheet.max_row + 1
    return empty_row

def append_new_record(sheet, empty_row, new_record_data,col_idx_value):
    # for col_idx, value in enumerate(new_record_data, start=1):
    for col_idx,value in col_idx_value:
        sheet.cell(row=empty_row, column=col_idx, value=value)




def run(col_idx_value):
    # Paths
    output_file_path = "output.xlsx"  # Adjust the output file name as needed
    template_file_path = "temp/temp.xlsx"

    # Load or create the workbook
    existing_wb, existing_sheet = load_existing_workbook(output_file_path)
    if existing_wb is None:
        existing_wb = load_workbook(template_file_path)
        existing_sheet = existing_wb.active

    # Data for the new record
    new_record_data = [
        "Business B", "License Class B", "Class B Name",  # Add more data fields as needed
    ]

    # Find an empty row to append the new record
    empty_row = find_empty_row(existing_sheet)

    # Append the new record data to the empty row
    append_new_record(existing_sheet, empty_row, new_record_data,col_idx_value)

    # Save the modified workbook to the output file
    existing_wb.save(output_file_path)

    print(f"New record appended to row {empty_row} in '{output_file_path}'")

# if __name__ == "__main__":
#     col_idx_value = [(1,"name"),(3,"data"),(2,"dd")]
#     run(col_idx_value)

    ############
    ############