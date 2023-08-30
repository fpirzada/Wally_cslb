

import openpyxl
from openpyxl.styles import Alignment
import os
import datetime

def create_or_load_workbook(filename):
    if os.path.exists(filename):
        return openpyxl.load_workbook(filename)
    else:
        workbook = openpyxl.Workbook()
        workbook.save(filename)
        return openpyxl.load_workbook(filename)

def format_header_and_subheader(worksheet, header_subheader_mapping):
    col_offset = 1
    for top_header, subheaders in header_subheader_mapping.items():
        col_num = col_offset
        worksheet.cell(row=1, column=col_num, value=top_header)
        worksheet.merge_cells(start_row=1, start_column=col_num, end_row=1, end_column=col_num + len(subheaders) - 1)
        worksheet.cell(row=1, column=col_num).alignment = Alignment(horizontal='center', vertical='center')

        for subheader in subheaders:
            worksheet.cell(row=2, column=col_num, value=subheader)
            col_num += 1

        col_offset += len(subheaders)

    # Insert an empty row as the 3rd row
    worksheet.insert_rows(3)


def find_empty_row(sheet):
    empty_row = None
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if all(cell is None for cell in row):
            empty_row = row_idx
            break

    if empty_row is None:
        empty_row = sheet.max_row + 1
    return empty_row

def append_live_data(worksheet,empty_row, live_data):
    for col_idx,value in live_data:
        # print(col_idx,value)
        worksheet.cell(row=empty_row, column=col_idx, value=value)


def auto_adjust_column_widths(worksheet):
    for col_idx, column_cells in enumerate(worksheet.columns, start=1):
        max_length = max(len(str(cell.value)) for cell in column_cells)
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        adjusted_width = max(max_length, len(col_letter)) + 2
        worksheet.column_dimensions[col_letter].width = adjusted_width

def save_workbook(workbook, filename):
    workbook.save(filename)

def main(live_data):
    current_date = datetime.datetime.now().strftime("%Y-%m-%d")
    filename = f"output_{current_date}.xlsx"
    
    workbook = create_or_load_workbook(filename)
    worksheet = workbook.active

    header_subheader_mapping = {
        "Business Information": ["Name", "Address","Phone #","City","Zip","License #", "Link", "Entity","Issue Date","Reissue Date","Expire Date"],
        "License Status": ["Status"],
        "Classifications": ["Classifications Name","Certifications Name"],
        "Workers' Compensation": ["Title","Policy Number","Effective Date","Expire Date","Workers' Compensation History Link","Insurance Company Code",],
        "Liability Insurance Information": ["Title","Policy Number","Amount","Effective Date","Expiration Date"],
        "Miscellaneous Information": ["Title"],
        "Other": ["Title"],
        "Personnel": ["Name","Title","Association Date","Classification"],

    }

    format_header_and_subheader(worksheet, header_subheader_mapping)

    
    # Find an empty row to append the new record
    empty_row = find_empty_row(worksheet)

    append_live_data(worksheet, empty_row,live_data)

    auto_adjust_column_widths(worksheet)

    save_workbook(workbook, filename)

